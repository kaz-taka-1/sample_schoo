import openpyxl as excel
from openpyxl.styles import PatternFill
import datetime

in_file = 'date-sample.xlsx'
out_file = 'date-mark_olditems.xlsx'
limit = datetime.datetime(2020, 1, 1)

def date_mark(in_file, out_file):
    book = excel.load_workbook(in_file)
    for sheet in book.worksheets:
        for row in sheet.iter_rows(min_row=4):
            check_row(row)

    book.save(out_file)

def check_row(row):
    date = row[0].value
    if type(date) is not datetime.datetime:
        return
    if date < limit:
        red = PatternFill(
            fill_type='solid',
            fgColor='ff0000')
        for cell in row:
            cell.fill = red

if __name__=='__main__':
    date_mark(in_file, out_file)