import openpyxl as excel

in_file = 'name1.xlsx'
out_file ='name_split.xlsx'

book = excel.load_workbook(in_file)
sheet = book.worksheets[0]

out_book = excel.Workbook()
out_sheet = out_book.active

for row in sheet.iter_rows():
    name = row[0].value
    sei, na = name.split(' ')
    out_sheet.append([sei, na])

out_book.save(out_file)