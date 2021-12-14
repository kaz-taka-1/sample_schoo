import  openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "西暦"
sheet["B1"] = "和暦"

start_y = 1930
for i in range(1000):
    sei = str(start_y + i)
    wa = '=TEXT("{}/1/1","gge年")'.format(sei)
    sheet.cell(row=(2+i), column=1, value=sei+'年')
    sheet.cell(row=(2+i), column=2, value=wa)
    print(sei, "=", wa)

book.save("wareki2.xlsx")