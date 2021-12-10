import openpyxl as excel

book = excel.Workbook()
sheet = book.active
sheet["A1"] = "勤勉な人の計画は必ず成功する"
sheet.cell(row=2, column=1, value="猿の尻笑い")
cell = sheet.cell(row=3, column=1)
cell.value = "探すのに時があり諦めるのに時がある"

book.save("cell_w.xlsx")