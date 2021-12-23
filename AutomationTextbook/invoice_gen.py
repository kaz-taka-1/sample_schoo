template_file = 'invoice-template.xlsx'
save_file = 'invoice01.xlsx'

name = '田中一郎'
subject = '1月分のご請求'
items = [ # 内訳
    ['リンゴ', 5, 320],
    ['バナナ', 8, 210],
    ['メロン', 5, 1200]
 ]

import openpyxl as excel
book = excel.load_workbook(template_file)
sheet = book.active
sheet["B4"] = name
sheet["C10"] = subject
total = 0
for i, it in enumerate(items):
    summary, count, price = it
    subtotal = count * price
    total += subtotal
    row = 15 + i
    sheet.cell(row, 2, summary)
    sheet.cell(row, 5, count)
    sheet.cell(row, 6, price)
    sheet.cell(row, 7, subtotal)

sheet["C1"] = total

book.save(save_file)