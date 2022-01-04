import openpyxl as excel, json

in_file = 'matome.json'
out_dir = './invoice02'
template_file = 'invoice-template.xlsx'
subject = "2月分のご請求"

def gen_invoice():
    with open(in_file, "rt") as fp:
        users = json.load(fp)
    for name, data in users.items():
        make_user_invoice(name, data)

def make_user_invoice(name, data):
    book = excel.load_workbook(template_file)
    sheet = book.active
    sheet["B4"] = name
    sheet["C10"] = subject
    sheet["C11"] = data["total"]

    for i, it in enumerate(data['items']):
        data, summary, cnt, price = it
        row = 15 + i
        sheet.cell(row, 2, summary+'('+data+')')
        sheet.cell(row, 5, cnt)
        sheet.cell(row, 6, price)
        sheet.cell(row, 7, cnt*price)

        out_file = out_dir+'/' + name + '様.xlsx'
        book.save(out_file)
        print("save:", out_file)

if __name__=="__main__":
    gen_invoice()