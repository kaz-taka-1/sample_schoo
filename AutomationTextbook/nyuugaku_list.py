import openpyxl as excel
import  datetime

book = excel.Workbook()
sheet = book.active

base_year = datetime.date.today().year - 10

for i in range(50):
    year = base_year + i
    s1 = year- 7
    s2 = year -6
    sf = "{}年4/2〜{}年4/1に生まれた人".format(s1, s2)
    sheet.cell(i+1, 1, value=str(year)+"年度")
    sheet.cell(i+1, 2, value=sf)

book.save("nyugaku_list.xlsx")